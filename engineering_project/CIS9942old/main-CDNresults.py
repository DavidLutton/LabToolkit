from CIS9942toXLSX import CIS9942toXLSX
from CIS9942toH import CIS9942toH
from CIS9942totxtHeader import CIS9942totxtHeader

import os

'''for root, directories, filenames in os.walk(b'.\E\.'):
    for directory in directories:
        True
        # print( os.path.join(root, directory) )
    for filename in filenames:
        file = os.path.join(root, filename)
        filepath = str(os.path.abspath(file), 'utf-8')
        if not filename.startswith(b'conv') and not filename.startswith(b'LPE334') and file.endswith(b'LC.RES'):
            print(filepath)
            CIS9942toXLSX(filepath, filename)
        if not filename.startswith(b'conv') and not filename.startswith(b'LPE334') and file.endswith(b'HC.RES'):
            print(filepath)
            CIS9942toXLSX(filepath, filename)

'''

'''
            CIS9942toXLSX(filepath, filename)

            CIS9942toH(filepath, filename)

            CIS9942totxtHeader(filepath, filename)
'''
# input("Finished")

from openpyxl import load_workbook


for root, directories, filenames in os.walk(b'.\E\.'):
    for directory in directories:
        True
        # print( os.path.join(root, directory) )
    for filename in filenames:
        file = os.path.join(root, filename)
        filepath = str(os.path.abspath(file), 'utf-8')
        if not filename.startswith(b'conv') and file.endswith(b'C.RES.xlsx'):
            print()
            #print(filepath)
            # CIS9942toXLSX(filepath, filename)
            #wb = load_workbook(filepath)
            
            lens = len(filepath) - 10
            print(filepath[:lens] + 'R' + filepath[lens+1:])
            #ws = wb.active
            #print(ws['A10'].value)
        
