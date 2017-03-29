import os
import _CIS9942
from openpyxl import Workbook
from openpyxl.chart import Reference, Series, LineChart

def CDN(YYYY, IDN, path):
    YYYY = str.encode(str(YYYY))
    datadir = b'Z:\\Calibration data\\' + path + b'\\' + IDN + b'\\' + YYYY + b' cal'
    print(datadir.decode())

    # wb = Workbook(guess_types=True)
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Calibration"
    ws2 = wb.create_sheet(title="Compression")
    ws3 = wb.create_sheet(title="Generator lvl")
    
    runcalibration = datadir + b'\\' + IDN + b'HR.RES'
    
    header, data, columns = _CIS9942.results(runcalibration)
    data, columns = _CIS9942.harmonise(data, columns)
    data, columns = _CIS9942.ordercolumns(data, columns)


    ws1.append(columns)
    for row in data:
        ws1.append(row)
    
    runcompression = datadir + b'\\' + IDN + b'HC.RES'
    
    header, data, columns = _CIS9942.results(runcompression)
    data, columns = _CIS9942.harmonise(data, columns)
    data, columns = _CIS9942.ordercolumns(data, columns)
    
    ws2.append(columns)
    for row in data:
        ws2.append(row)
    
    points = len(data)


    ws3.append(["Frequency", "Calibration", "Compression", "", "Min", "Result", "Max"])
    
    for x in range(1, points + 1):
        y = str(x + 1)
        ws3.append(["=Calibration!A" + y, "=Calibration!C" + y , "=Compression!C" + y, "", 3.1, "=C" + y + "-B" + y , 7.1])
        
    chart = LineChart()
    
    labels = Reference(ws3, min_col=1, min_row=5, max_row=7)
    x = Reference(ws3, min_col=1, min_row=2, max_row=points)
    
    # chart.series = ( Series(Reference(ws3, min_col=5, min_row=1, max_row=points), title_from_data=True),)
    chart.append(Series(Reference(ws3, min_col=5, min_row=1, max_row=points + 1), title_from_data=True))
    chart.append(Series(Reference(ws3, min_col=6, min_row=1, max_row=points + 1), title_from_data=True))
    chart.append(Series(Reference(ws3, min_col=7, min_row=1, max_row=points + 1), title_from_data=True))
    chart.title = IDN.decode() + " Compression"

    timings = Reference(ws3, min_row=2, max_row=points + 1, min_col=1, max_col=1)
    chart.set_categories(timings)

    chart.x_axis.title = 'MHz'
    chart.y_axis.title = 'dB'
    
    cs = wb.create_chartsheet()
    cs.add_chart(chart)
    #ws3.add_chart(chart, 'Z1')
        
    dest = datadir + b'\\' + IDN + b' compression.xlsx'
    wb.save(filename = dest.decode() )
    print(dest.decode())


def CDNl(YYYY, IDN, path):
    YYYY = str.encode(str(YYYY))
    datadir = b'Z:\\Calibration data\\' + path + b'\\' + IDN + b'\\' + YYYY + b' cal'
    print(datadir.decode())

    # wb = Workbook(guess_types=True)
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Calibration"
    ws2 = wb.create_sheet(title="Compression")
    ws3 = wb.create_sheet(title="Generator lvl")
    
    runcalibration = datadir + b'\\' + IDN + b'LR.RES'
    
    header, data, columns = _CIS9942.results(runcalibration)
    data, columns = _CIS9942.harmonise(data, columns)
    data, columns = _CIS9942.ordercolumns(data, columns)


    ws1.append(columns)
    for row in data:
        ws1.append(row)
    
    runcompression = datadir + b'\\' + IDN + b'LC.RES'
    
    header, data, columns = _CIS9942.results(runcompression)
    data, columns = _CIS9942.harmonise(data, columns)
    data, columns = _CIS9942.ordercolumns(data, columns)
    
    ws2.append(columns)
    for row in data:
        ws2.append(row)
    
    points = len(data)


    ws3.append(["Frequency", "Calibration", "Compression", "", "Min", "Result", "Max"])
    
    for x in range(1, points + 1):
        y = str(x + 1)
        ws3.append(["=Calibration!A" + y, "=Calibration!C" + y , "=Compression!C" + y, "", 3.1, "=C" + y + "-B" + y , 7.1])
        
    chart = LineChart()
    
    labels = Reference(ws3, min_col=1, min_row=5, max_row=7)
    x = Reference(ws3, min_col=1, min_row=2, max_row=points)
    
    # chart.series = ( Series(Reference(ws3, min_col=5, min_row=1, max_row=points), title_from_data=True),)
    chart.append(Series(Reference(ws3, min_col=5, min_row=1, max_row=points + 1), title_from_data=True))
    chart.append(Series(Reference(ws3, min_col=6, min_row=1, max_row=points + 1), title_from_data=True))
    chart.append(Series(Reference(ws3, min_col=7, min_row=1, max_row=points + 1), title_from_data=True))
    chart.title = IDN.decode() + " Compression"

    timings = Reference(ws3, min_row=2, max_row=points + 1, min_col=1, max_col=1)
    chart.set_categories(timings)

    chart.x_axis.title = 'MHz'
    chart.y_axis.title = 'dB'
    
    cs = wb.create_chartsheet()
    cs.add_chart(chart)
    #ws3.add_chart(chart, 'Z1')
        
    dest = datadir + b'\\' + IDN + b' compression.xlsx'
    wb.save(filename = dest.decode() )
    print(dest.decode())

